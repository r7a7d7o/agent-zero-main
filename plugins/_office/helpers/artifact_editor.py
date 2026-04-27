from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape
import xml.etree.ElementTree as ET

from plugins._office.helpers import wopi_store


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
P_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_NS = "http://www.w3.org/XML/1998/namespace"

for prefix, namespace in {
    "w": W_NS,
    "a": A_NS,
    "p": P_NS,
    "r": R_NS,
}.items():
    ET.register_namespace(prefix, namespace)


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def read_artifact(doc: dict[str, Any], max_chars: int = 12000) -> dict[str, Any]:
    """Extract compact editable content from an Office artifact."""
    path = Path(doc["path"])
    ext = str(doc["extension"]).lower()
    if ext == "docx":
        content = _read_docx(path)
    elif ext == "xlsx":
        content = _read_xlsx(path)
    elif ext == "pptx":
        content = _read_pptx(path)
    elif ext in {"odt", "ods", "odp"}:
        content = _read_odf(path)
    else:
        raise ValueError(f"Unsupported Office format: {ext}")

    return _trim_payload(content, max_chars=max_chars)


def edit_artifact(
    doc: dict[str, Any],
    operation: str = "",
    content: str = "",
    find: str = "",
    replace: str = "",
    sheet: str = "",
    cells: Any = None,
    rows: Any = None,
    slides: Any = None,
    **kwargs: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Apply a direct saved edit to an Office artifact and return updated metadata."""
    path = Path(doc["path"])
    ext = str(doc["extension"]).lower()
    op = normalize_operation(operation, content=content, find=find, cells=cells, rows=rows, slides=slides)
    before = path.read_bytes()

    if ext == "docx":
        updated, details = _edit_docx(before, op, content=content, find=find, replace=replace, **kwargs)
    elif ext == "xlsx":
        updated, details = _edit_xlsx(path, op, content=content, find=find, replace=replace, sheet=sheet, cells=cells, rows=rows, **kwargs)
    elif ext == "pptx":
        updated, details = _edit_pptx(before, op, content=content, find=find, replace=replace, slides=slides, **kwargs)
    else:
        raise ValueError(f"Direct edit is not available for .{ext}. Use Collabora in the Office canvas.")

    changed = updated != before
    updated_doc = (
        wopi_store.replace_document_bytes(doc["file_id"], updated, actor="document_artifact:edit")
        if changed
        else doc
    )
    preview = read_artifact(updated_doc, max_chars=int(kwargs.get("preview_chars") or 4000))
    payload = {
        "ok": True,
        "action": "edit",
        "operation": op,
        "changed": changed,
        **details,
        "preview": preview,
    }
    return updated_doc, payload


def normalize_operation(
    operation: str,
    *,
    content: str = "",
    find: str = "",
    cells: Any = None,
    rows: Any = None,
    slides: Any = None,
) -> str:
    op = str(operation or "").strip().lower().replace("-", "_")
    aliases = {
        "patch": "replace_text" if find else "set_text",
        "update": "replace_text" if find else "set_text",
        "replace": "replace_text",
        "append": "append_text",
        "prepend": "prepend_text",
        "write": "set_text",
        "set": "set_text",
        "set_content": "set_text",
        "set_sheet": "set_rows",
        "write_sheet": "set_rows",
        "add_rows": "append_rows",
        "add_slide": "append_slide",
        "set_deck": "set_slides",
    }
    op = aliases.get(op, op)
    if op:
        return op
    if cells:
        return "set_cells"
    if rows:
        return "append_rows"
    if slides:
        return "set_slides"
    if find:
        return "replace_text"
    if content:
        return "set_text"
    raise ValueError("operation is required")


def _read_docx(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs = []
    for paragraph in root.iter(qn(W_NS, "p")):
        text = "".join(node.text or "" for node in paragraph.iter(qn(W_NS, "t")))
        if text.strip():
            paragraphs.append(text)
    return {
        "kind": "document",
        "paragraph_count": len(paragraphs),
        "text": "\n".join(paragraphs),
        "paragraphs": paragraphs[:80],
    }


def _read_xlsx(path: Path) -> dict[str, Any]:
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(path, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets[:8]:
        rows = []
        max_row = min(worksheet.max_row or 0, 80)
        max_col = min(worksheet.max_column or 0, 30)
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            values = ["" if value is None else value for value in row]
            if any(str(value).strip() for value in values):
                rows.append(values)
        sheets.append({
            "name": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "preview_rows": rows,
        })
    return {
        "kind": "spreadsheet",
        "sheet_count": len(workbook.worksheets),
        "sheets": sheets,
    }


def _read_pptx(path: Path) -> dict[str, Any]:
    slides = []
    with zipfile.ZipFile(path) as archive:
        for name in _slide_names(archive):
            root = ET.fromstring(archive.read(name))
            lines = []
            for paragraph in root.iter(qn(A_NS, "p")):
                text = "".join(node.text or "" for node in paragraph.iter(qn(A_NS, "t"))).strip()
                if text:
                    lines.append(text)
            slides.append({
                "index": len(slides) + 1,
                "title": lines[0] if lines else "",
                "lines": lines,
            })
    return {
        "kind": "presentation",
        "slide_count": len(slides),
        "slides": slides[:40],
    }


def _read_odf(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("content.xml")
    root = ET.fromstring(xml)
    text = "\n".join((node.text or "").strip() for node in root.iter() if (node.text or "").strip())
    return {
        "kind": "office_document",
        "text": text,
    }


def _edit_docx(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "append_text", "prepend_text", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported DOCX operation: {op}")

    with zipfile.ZipFile(io.BytesIO(before)) as archive:
        files = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    root = ET.fromstring(files["word/document.xml"])

    if op == "replace_text" or op == "delete_text":
        if not find:
            raise ValueError("find is required for replace_text")
        replacement = "" if op == "delete_text" else replace
        count = _replace_text_in_paragraphs(
            root,
            paragraph_tag=qn(W_NS, "p"),
            text_tag=qn(W_NS, "t"),
            set_text=_set_word_paragraph_text,
            find=find,
            replacement=replacement,
            limit=_int_or_none(kwargs.get("count")),
        )
        details = {"replacements": count}
        if count == 0:
            return before, details
    else:
        lines = _text_lines(content)
        body = root.find(f".//{qn(W_NS, 'body')}")
        if body is None:
            raise ValueError("DOCX document body not found")
        paragraphs = [_word_paragraph(line) for line in lines]
        if op == "set_text":
            sect_pr = [child for child in list(body) if child.tag == qn(W_NS, "sectPr")]
            for child in list(body):
                body.remove(child)
            for paragraph in paragraphs:
                body.append(paragraph)
            for child in sect_pr:
                body.append(child)
        elif op == "append_text":
            insert_at = len(body)
            for idx, child in enumerate(list(body)):
                if child.tag == qn(W_NS, "sectPr"):
                    insert_at = idx
                    break
            for paragraph in reversed(paragraphs):
                body.insert(insert_at, paragraph)
        elif op == "prepend_text":
            for paragraph in reversed(paragraphs):
                body.insert(0, paragraph)
        details = {"paragraphs_written": len(paragraphs)}

    files["word/document.xml"] = _xml_bytes(root)
    return _zip_from_existing(files), details


def _edit_xlsx(
    path: Path,
    op: str,
    *,
    content: str = "",
    find: str = "",
    replace: str = "",
    sheet: str = "",
    cells: Any = None,
    rows: Any = None,
    **kwargs: Any,
) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "set_rows", "append_text", "append_rows", "set_cells", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported XLSX operation: {op}")
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(path)
    worksheet = _worksheet(workbook, sheet)

    details: dict[str, Any] = {"sheet": worksheet.title}
    if op in {"set_text", "set_rows"}:
        parsed_rows = _normalize_rows(rows if rows is not None else content)
        _clear_worksheet(worksheet)
        _write_rows(worksheet, parsed_rows, start_row=1)
        details["rows_written"] = len(parsed_rows)
    elif op in {"append_text", "append_rows"}:
        parsed_rows = _normalize_rows(rows if rows is not None else content)
        start_row = max((worksheet.max_row or 0) + 1, 1)
        _write_rows(worksheet, parsed_rows, start_row=start_row)
        details["rows_appended"] = len(parsed_rows)
        details["start_row"] = start_row
    elif op == "set_cells":
        assignments = _normalize_cells(cells, default_sheet=worksheet.title)
        for sheet_name, cell, value in assignments:
            target = _worksheet(workbook, sheet_name)
            target[cell] = value
        details["cells_written"] = len(assignments)
    elif op in {"replace_text", "delete_text"}:
        if not find:
            raise ValueError("find is required for replace_text")
        replacement = "" if op == "delete_text" else replace
        count = 0
        limit = _int_or_none(kwargs.get("count"))
        for target in workbook.worksheets:
            for row in target.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or find not in cell.value:
                        continue
                    remaining = None if limit is None else max(limit - count, 0)
                    if remaining == 0:
                        break
                    cell.value, replaced = _replace_limited(cell.value, find, replacement, remaining)
                    count += replaced
                if limit is not None and count >= limit:
                    break
            if limit is not None and count >= limit:
                break
        details["replacements"] = count
        if count == 0:
            return path.read_bytes(), details

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), details


def _edit_pptx(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", slides: Any = None, **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "set_slides", "append_text", "append_slide", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported PPTX operation: {op}")

    if op in {"set_text", "set_slides"}:
        parsed_slides = _normalize_slides(slides if slides is not None else content)
        return _pptx_from_slides(parsed_slides), {"slides_written": len(parsed_slides)}

    if op in {"append_text", "append_slide"}:
        existing = _pptx_text_slides(before)
        existing.extend(_normalize_slides(slides if slides is not None else content))
        return _pptx_from_slides(existing), {"slides_written": len(existing)}

    with zipfile.ZipFile(io.BytesIO(before)) as archive:
        files = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    if not find:
        raise ValueError("find is required for replace_text")
    replacement = "" if op == "delete_text" else replace
    count = 0
    limit = _int_or_none(kwargs.get("count"))
    for name in sorted([name for name in files if name.startswith("ppt/slides/slide") and name.endswith(".xml")], key=_natural_key):
        root = ET.fromstring(files[name])
        count += _replace_text_in_paragraphs(
            root,
            paragraph_tag=qn(A_NS, "p"),
            text_tag=qn(A_NS, "t"),
            set_text=_set_drawing_paragraph_text,
            find=find,
            replacement=replacement,
            limit=None if limit is None else max(limit - count, 0),
        )
        files[name] = _xml_bytes(root)
        if limit is not None and count >= limit:
            break
    if count == 0:
        return before, {"replacements": count}
    return _zip_from_existing(files), {"replacements": count}


def _replace_text_in_paragraphs(
    root: ET.Element,
    *,
    paragraph_tag: str,
    text_tag: str,
    set_text: Any,
    find: str,
    replacement: str,
    limit: int | None,
) -> int:
    count = 0
    for paragraph in root.iter(paragraph_tag):
        texts = list(paragraph.iter(text_tag))
        if not texts:
            continue
        current = "".join(node.text or "" for node in texts)
        if find not in current:
            continue
        remaining = None if limit is None else max(limit - count, 0)
        if remaining == 0:
            break
        updated, replaced = _replace_limited(current, find, replacement, remaining)
        if replaced:
            set_text(paragraph, updated)
            count += replaced
    return count


def _replace_limited(value: str, find: str, replacement: str, limit: int | None) -> tuple[str, int]:
    if limit is None:
        return value.replace(find, replacement), value.count(find)
    return value.replace(find, replacement, limit), min(value.count(find), limit)


def _set_word_paragraph_text(paragraph: ET.Element, text: str) -> None:
    keep = [child for child in list(paragraph) if child.tag == qn(W_NS, "pPr")]
    for child in list(paragraph):
        paragraph.remove(child)
    for child in keep:
        paragraph.append(child)
    paragraph.append(_word_run(text))


def _word_paragraph(text: str) -> ET.Element:
    paragraph = ET.Element(qn(W_NS, "p"))
    paragraph.append(_word_run(text))
    return paragraph


def _word_run(text: str) -> ET.Element:
    run = ET.Element(qn(W_NS, "r"))
    text_node = ET.SubElement(run, qn(W_NS, "t"))
    if text.startswith(" ") or text.endswith(" "):
        text_node.set(qn(XML_NS, "space"), "preserve")
    text_node.text = text
    return run


def _set_drawing_paragraph_text(paragraph: ET.Element, text: str) -> None:
    keep = [child for child in list(paragraph) if child.tag == qn(A_NS, "pPr")]
    for child in list(paragraph):
        paragraph.remove(child)
    for child in keep:
        paragraph.append(child)
    run = ET.SubElement(paragraph, qn(A_NS, "r"))
    text_node = ET.SubElement(run, qn(A_NS, "t"))
    text_node.text = text


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for spreadsheet edits") from exc
    return openpyxl


def _worksheet(workbook: Any, sheet: str = "") -> Any:
    if sheet:
        if sheet not in workbook.sheetnames:
            return workbook.create_sheet(sheet)
        return workbook[sheet]
    return workbook.active


def _clear_worksheet(worksheet: Any) -> None:
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)


def _write_rows(worksheet: Any, rows: list[list[Any]], start_row: int) -> None:
    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):
            worksheet.cell(row=start_row + row_offset, column=1 + col_offset, value=_cell_value(value))


def _normalize_rows(value: Any) -> list[list[Any]]:
    if value is None:
        return []
    if isinstance(value, list):
        rows = value
    elif isinstance(value, str):
        rows = _rows_from_text(value)
    else:
        rows = [[value]]
    normalized = []
    for row in rows:
        if isinstance(row, (list, tuple)):
            normalized.append([_cell_value(value) for value in row])
        else:
            normalized.append([_cell_value(row)])
    return normalized


def _normalize_cells(cells: Any, default_sheet: str) -> list[tuple[str, str, Any]]:
    if isinstance(cells, str):
        parsed = json.loads(cells)
    else:
        parsed = cells
    if not parsed:
        raise ValueError("cells is required for set_cells")

    result: list[tuple[str, str, Any]] = []
    if isinstance(parsed, dict):
        for ref, value in parsed.items():
            sheet, cell = _split_cell_ref(str(ref), default_sheet)
            result.append((sheet, cell, _cell_value(value)))
    elif isinstance(parsed, list):
        for item in parsed:
            if not isinstance(item, dict):
                raise ValueError("cells list entries must be objects")
            ref = str(item.get("cell") or item.get("ref") or "")
            sheet = str(item.get("sheet") or default_sheet)
            if "!" in ref:
                sheet, ref = _split_cell_ref(ref, default_sheet)
            if not ref:
                raise ValueError("cell is required for each cells entry")
            result.append((sheet, ref, _cell_value(item.get("value"))))
    else:
        raise ValueError("cells must be an object or list")
    return result


def _split_cell_ref(ref: str, default_sheet: str) -> tuple[str, str]:
    if "!" not in ref:
        return default_sheet, ref
    sheet, cell = ref.split("!", 1)
    return sheet.strip("'") or default_sheet, cell


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _rows_from_text(content: str) -> list[list[str]]:
    text = str(content or "").strip("\n")
    if not text.strip():
        return []
    lines = [line for line in text.splitlines() if line.strip()]
    markdown_rows = _markdown_table_rows(lines)
    if markdown_rows:
        return markdown_rows

    delimiter = "\t" if any("\t" in line for line in lines) else ("," if any("," in line for line in lines) else None)
    if delimiter:
        return [row for row in csv.reader(io.StringIO("\n".join(lines)), delimiter=delimiter)]
    return [[line] for line in lines]


def _markdown_table_rows(lines: list[str]) -> list[list[str]]:
    table_lines = [line.strip() for line in lines if line.strip().startswith("|") and line.strip().endswith("|")]
    if len(table_lines) < 2:
        return []
    rows = []
    for line in table_lines:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if all(re.fullmatch(r":?-{3,}:?", cell or "") for cell in cells):
            continue
        rows.append(cells)
    return rows


def _pptx_text_slides(data: bytes) -> list[dict[str, Any]]:
    slides = []
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        for name in _slide_names(archive):
            root = ET.fromstring(archive.read(name))
            lines = []
            for paragraph in root.iter(qn(A_NS, "p")):
                text = "".join(node.text or "" for node in paragraph.iter(qn(A_NS, "t"))).strip()
                if text:
                    lines.append(text)
            if lines:
                slides.append({"title": lines[0], "bullets": lines[1:]})
    return slides


def _normalize_slides(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        if stripped.startswith("[") or stripped.startswith("{"):
            return _normalize_slides(json.loads(stripped))
        chunks = re.split(r"(?m)^\s*---+\s*$", stripped)
        result = []
        for chunk in chunks:
            lines = [line.strip(" -\t") for line in chunk.splitlines() if line.strip()]
            if not lines:
                continue
            result.append({"title": lines[0], "bullets": lines[1:]})
        return result
    if isinstance(value, dict):
        return [_slide_from_mapping(value)]
    if isinstance(value, list):
        result = []
        for item in value:
            if isinstance(item, dict):
                result.append(_slide_from_mapping(item))
            elif isinstance(item, str):
                result.extend(_normalize_slides(item))
            elif isinstance(item, (list, tuple)):
                lines = [str(part) for part in item if str(part).strip()]
                if lines:
                    result.append({"title": lines[0], "bullets": lines[1:]})
            else:
                result.append({"title": str(item), "bullets": []})
        return result
    return [{"title": str(value), "bullets": []}]


def _slide_from_mapping(value: dict[str, Any]) -> dict[str, Any]:
    title = str(value.get("title") or value.get("heading") or "Slide")
    bullets = value.get("bullets")
    if bullets is None:
        body = value.get("body") or value.get("content") or ""
        bullets = [line.strip(" -\t") for line in str(body).splitlines() if line.strip()]
    elif isinstance(bullets, str):
        bullets = [line.strip(" -\t") for line in bullets.splitlines() if line.strip()]
    else:
        bullets = [str(item) for item in bullets]
    return {"title": title, "bullets": bullets}


def _pptx_from_slides(slides: list[dict[str, Any]]) -> bytes:
    if not slides:
        slides = [{"title": "Presentation", "bullets": []}]

    files: dict[str, str | bytes] = {
        "[Content_Types].xml": _pptx_content_types(len(slides)),
        "_rels/.rels": (
            '<?xml version="1.0" encoding="UTF-8"?>'
            f'<Relationships xmlns="{REL_NS}">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="ppt/presentation.xml"/>'
            "</Relationships>"
        ),
        "ppt/_rels/presentation.xml.rels": _pptx_presentation_rels(len(slides)),
        "ppt/presentation.xml": _pptx_presentation_xml(len(slides)),
    }
    for index, slide in enumerate(slides, start=1):
        files[f"ppt/slides/slide{index}.xml"] = _pptx_slide_xml(slide)
    return _zip_map(files)


def _pptx_content_types(count: int) -> str:
    overrides = [
        '<Override PartName="/ppt/presentation.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>'
    ]
    for index in range(1, count + 1):
        overrides.append(
            f'<Override PartName="/ppt/slides/slide{index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<Types xmlns="{CT_NS}">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        + "".join(overrides)
        + "</Types>"
    )


def _pptx_presentation_rels(count: int) -> str:
    rels = []
    for index in range(1, count + 1):
        rels.append(
            f'<Relationship Id="rId{index}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" '
            f'Target="slides/slide{index}.xml"/>'
        )
    return '<?xml version="1.0" encoding="UTF-8"?>' + f'<Relationships xmlns="{REL_NS}">' + "".join(rels) + "</Relationships>"


def _pptx_presentation_xml(count: int) -> str:
    slide_ids = "".join(f'<p:sldId id="{255 + index}" r:id="rId{index}"/>' for index in range(1, count + 1))
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<p:presentation xmlns:p="{P_NS}" xmlns:r="{R_NS}">'
        f"<p:sldIdLst>{slide_ids}</p:sldIdLst>"
        '<p:sldSz cx="9144000" cy="5143500"/>'
        "</p:presentation>"
    )


def _pptx_slide_xml(slide: dict[str, Any]) -> str:
    title = str(slide.get("title") or "Slide")
    bullets = [str(item) for item in slide.get("bullets") or []]
    paragraphs = [title, *bullets]
    text = "".join(f"<a:p><a:r><a:t>{escape(item)}</a:t></a:r></a:p>" for item in paragraphs)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<p:sld xmlns:a="{A_NS}" xmlns:p="{P_NS}">'
        "<p:cSld><p:spTree>"
        '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
        "<p:grpSpPr/>"
        '<p:sp><p:nvSpPr><p:cNvPr id="2" name="Content"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>'
        f"<p:txBody><a:bodyPr/><a:lstStyle/>{text}</p:txBody>"
        "</p:sp>"
        "</p:spTree></p:cSld>"
        "</p:sld>"
    )


def _slide_names(archive: zipfile.ZipFile) -> list[str]:
    return sorted(
        [name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")],
        key=_natural_key,
    )


def _natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", value)]


def _text_lines(content: str) -> list[str]:
    lines = [line.rstrip() for line in str(content or "").splitlines()]
    return lines or [""]


def _int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def _xml_bytes(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _zip_from_existing(files: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, data in files.items():
            archive.writestr(name, data)
    return buffer.getvalue()


def _zip_map(files: dict[str, str | bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, value in files.items():
            archive.writestr(name, value.encode("utf-8") if isinstance(value, str) else value)
    return buffer.getvalue()


def _trim_payload(payload: dict[str, Any], max_chars: int) -> dict[str, Any]:
    text = json.dumps(payload, ensure_ascii=False, default=str)
    if len(text) <= max_chars:
        return payload
    trimmed = dict(payload)
    if "paragraphs" in trimmed:
        trimmed["paragraphs"] = trimmed["paragraphs"][:20]
    if "sheets" in trimmed:
        trimmed["sheets"] = [
            {**sheet, "preview_rows": sheet.get("preview_rows", [])[:20]}
            for sheet in trimmed["sheets"][:4]
        ]
    if "slides" in trimmed:
        trimmed["slides"] = trimmed["slides"][:12]
    if "text" in trimmed and isinstance(trimmed["text"], str):
        trimmed["text"] = trimmed["text"][:max_chars] + "\n... [truncated]"
    trimmed["truncated"] = True
    return trimmed
